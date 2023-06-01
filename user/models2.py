from django.db import models
from django.contrib.auth.models import User
from django.dispatch import receiver
from django.db.models.signals import post_save
from django.db.models.signals import pre_delete
from django import forms
from django.forms.widgets import CheckboxSelectMultiple
import openpyxl
from datetime import datetime
from dateutil.parser import parse
from django.db.models.signals import post_delete

import requests
from io import BytesIO

import os
from PIL import Image
import pyautogui
from django.conf import settings
import asyncio
from urllib.parse import urlparse
import openpyxl
from dateutil.parser import parse
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from pyppeteer import launch


class tag(models.Model):
    tag = models.CharField(max_length=50)

    def __str__(self):
        return self.tag

class event(models.Model):
    categories=(("GD","GD"),("BD","BD"),("GD/BD","GD/BD"),("LGD","LGD"),("LBD","LBD"),("GrD","GrD"))
    category = models.CharField(max_length=200,choices=categories,null=True)
    author = models.ForeignKey(User, on_delete=models.CASCADE,default=User)
    event_date = models.DateField()
    event_title= models.CharField(max_length=200)
    event_bullet_1=models.CharField(max_length=200,help_text = "Required",blank=False,null=True)
    event_bullet_2=models.CharField(max_length=200,help_text = "Opional",null=True,blank=True)
    event_bullet_3=models.CharField(max_length=200,help_text = "Opional",null=True,blank=True)
    event_bullet_4=models.CharField(max_length=200,help_text = "Opional",null=True,blank=True)
    event_bullet_5=models.CharField(max_length=200,help_text = "Opional",null=True,blank=True)
    source1_url = models.URLField(help_text = "Required",)
    source1_image = models.URLField(help_text = "Required",null=True,blank=True)
    source2_url = models.URLField(help_text = "Opional",null=True,blank=True)
    source2_image = models.URLField(help_text = "Opional",null=True,blank=True)
    source3_url = models.URLField(help_text = "Opional",null=True,blank=True)
    source3_image = models.URLField(help_text = "Opional",null=True,blank=True)
    tags = models.ManyToManyField("tag",null=True,blank=True)
    def __str__(self):
        return self.event_title
    
class excel_file(models.Model):
    author = models.ForeignKey(User, on_delete=models.CASCADE,default=User)
    file = models.FileField(help_text="up to 20 reports",blank=True)
    def __str__(self):
        return self.author.username
    
@receiver(post_save, sender=excel_file)
def do_something_with_file(sender, instance, created, **kwargs):
    if created:
        create_events_from_excel(instance.file.path)



def capture_screenshot(url):
    response = requests.get(url)
    img = Image.open(BytesIO(response.content))
    return img


def create_events_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    row_num=2
    for row in ws.iter_rows(min_row=2, values_only=True):  
        CATEGORY, DATE, TITLE,BULLET1, BULLET2, BULLET3,BULLET4, BULLET5, SOURCE1, SOURCE2,SOURCE3, TAGS = row
        source1_url = ws.cell(row=row_num, column=9).value
        source2_url = ws.cell(row=row_num, column=10).value
        source3_url = ws.cell(row=row_num, column=11).value
        TAGS = TAGS.split(',')
        tags_list = []
        for name in TAGS:
            tat_el, created = tag.objects.get_or_create(tag=name.strip())
            tags_list.append(tat_el)
        print(tags_list)
        tags_list = set(tags_list)
        print(tags_list)
        if source1_url is not None:
            hyperlink = ws.cell(row=row_num, column=9).hyperlink
            if hyperlink is not None:
                source1_url = hyperlink.target
        if source2_url != None:
            source2_url = ws.cell(row=row_num, column=10).hyperlink.target
        if source3_url != None:
            source3_url = ws.cell(row=row_num, column=11).hyperlink.target
        row_num+=1
        # screenshot = capture_screenshot(source1_url)


        try:
            date_obj = parse(str(DATE))
            new_date_str = date_obj.strftime('%Y-%m-%d')

            filename = os.path.basename(urlparse(source1_url).path)
            file_path = os.path.join(settings.MEDIA_ROOT, filename)
            asyncio.get_event_loop().run_until_complete(capture_screenshot(source1_url, file_path))
            
            with open(file_path, 'rb') as f:
                content = ContentFile(f.read())


            event_obj =event.objects.create(author=User.objects.get_or_create(username='admin')[0],category=CATEGORY, event_date=new_date_str, event_title=TITLE,event_bullet_1=BULLET1, event_bullet_2=BULLET2, event_bullet_3=BULLET3,event_bullet_4=BULLET4, event_bullet_5=BULLET5, source1_url=source1_url, source2_url=source2_url, source3_url=source3_url,source1_image=default_storage.save(filename, content, save=False))
            event_obj.tags.set(tags_list)
        except ValueError:
            pass


async def capture_screenshot(url, file_path):
    browser = await launch()
    page = await browser.newPage()
    await page.goto(url)
    await page.screenshot(path=file_path)
    await browser.close()